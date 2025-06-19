#!/usr/bin/env python
# coding: utf-8
import os
import time
from datetime import datetime
import shutil
import openpyxl
import traceback
import logging
from datetime import datetime
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from concurrent.futures import ThreadPoolExecutor
from path_config import *
from PREPROCESSING.preprocessing_black_cell import *
from PREPROCESSING.preprocessing_id import *
from PREPROCESSING.preprocessing_printed import *
from PREPROCESSING.preprocessing_cropped_rfg import *
from PREPROCESSING.preprocessing_cropped_cnn import *
from PREPROCESSING.preprocessing_cropped_dnn import *
from PREPROCESSING.preprocessing_cropped_ann import *
from OCR_PROGRAM.train_model_handw_gemini import *
from OCR_PROGRAM.load_json import *
from OCR_PROGRAM.OCR_id import *
from OCR_PROGRAM.OCR_ann import *
from OCR_PROGRAM.OCR_cnn import *
from OCR_PROGRAM.OCR_rfg import *
from OCR_PROGRAM.OCR_dnn import *
from model_ensemble import *
from save_to_excel import *
from OCR_PROGRAM.OCR_resnet18 import *
from Result_analysis import *

class FileProcessor:
    def __init__(self):
        self.currently_processing = False
        self.processed_files = []
        self.wb_target = None
        self.excel_path = None
        self.observer = None
        self.executor = None

        self.initialize_system()
        
    def initialize_system(self):
        # Set up logger
        log_file = LOGGING
        logging.basicConfig(
            filename=log_file,
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )

        logger = logging.getLogger(__name__)

        logger.info("Loading path env...")
        self.scan_folder = SCAN_FOLDER_PATH
        current_date = datetime.today().strftime("%Y-%m-%d_%H-%M-%S")
        
        self.temp_folder = TEMP_PATH
        self.excel_temp_folder = os.path.join(self.temp_folder, "TEMP_EXCEL/")

        def clear_directory(directory_path):
            if not os.path.exists(directory_path):
                print(f"Error: Directory '{directory_path}' does not exist.")
                return
            
            if not os.path.isdir(directory_path):
                print(f"Error: Path '{directory_path}' is not a directory.")
                return

            print(f"Clearing contents of: {directory_path}")
            for item in os.listdir(directory_path):
                item_path = os.path.join(directory_path, item)
                try:
                    if os.path.isfile(item_path) or os.path.islink(item_path):
                        os.unlink(item_path)
                        print(f"Removed file: {item_path}")
                except Exception as e:
                    print(f"Failed to delete {item_path}. Reason: {e}")
            print(f"Directory '{directory_path}' cleared.")
        
        clear_directory(self.excel_temp_folder)

        self.excel_path = os.path.join(self.excel_temp_folder, f"data_{current_date}.xlsx")
        logger.info("File created")
        print(self.excel_path)
        print(EXCEL_TEMPLATE_PATH)
        print(os.path.exists(EXCEL_TEMPLATE_PATH))
        if not os.path.exists(self.excel_path):
            logger.info("Copying excel template...")
            shutil.copyfile(EXCEL_TEMPLATE_PATH, self.excel_path)

        logger.info("Loading excel workbook...")
        start_load = time.time()
        self.wb_target = openpyxl.load_workbook(self.excel_path)
        end_load = time.time()
        load_time = end_load - start_load
        logger.info(f"Excel workbook has been loaded. Total time: {load_time:.2f} seconds.")

    def process_new_file(self, image_path):
        """Process a new image file."""
        self.currently_processing = True
        logger.info(f"Processing new file: {image_path}")
        
        try:
            # Load and preprocess image
            image = load_image(image_path)
            processed_image = preprocessing_image(image, image_path)
            processed_img_id = preprocessing_id(image, image_path)
            # Load ROI (Region of Interest)
            roi_id_scan, loaded_roi_id_scan = load_json_scan()
            # Do OCR for ID
            text_id_scan, cropped = ocr_id_scan(processed_image, roi_id_scan, loaded_roi_id_scan)
            print(text_id_scan[0])

            if text_id_scan[0] == True:
                processed_black = preprocessing_black_cell(image)
                roi_id_checksheet, loaded_roi_id_checksheet = load_json_identifier()
                
                # Using model status to do OCR with activated the model
                if model_cnn_status:
                    proba_final_id_cnn, tc_result_id = ocr_id_cnn(processed_img_id, roi_id_checksheet, loaded_roi_id_checksheet)
                else:
                    proba_final_id_cnn, tc_result_id = {}, {}

                if model_dnn_status:
                    proba_final_id_dnn = ocr_id_dnn(processed_img_id, roi_id_checksheet, loaded_roi_id_checksheet)
                else:
                    proba_final_id_dnn = {}

                if model_ann_status:
                    proba_final_id_ann = ocr_id_ann(processed_img_id, roi_id_checksheet, loaded_roi_id_checksheet)
                else:
                    proba_final_id_ann = {}

                if model_rfg_status:
                    proba_final_id_rfg = ocr_id_rfg(processed_img_id, roi_id_checksheet, loaded_roi_id_checksheet)
                else:
                    proba_final_id_rfg = {}

            
                proba_idc, tc_result_idc, list_of_var_id = ensemble_model_digit_print(
                    proba_final_id_cnn, tc_result_id, proba_final_id_dnn,
                    proba_final_id_ann, proba_final_id_rfg
                )

                proba_idc = list_of_dicts_to_dict(proba_idc)
                tc_result_idc = list_of_dicts_to_dict(tc_result_idc)

                proba_id = proba_idc["B1"]
                final_confidence = tc_result_idc["CT1"]
                print(proba_id)
                print(final_confidence)

                # Allocate result based on checksheet type
                time_stamp = datetime.today().strftime("%H-%M-%S")

                logger.info(f"Classifying the image with final ID: {proba_id}")
                ocr_prompts = load_prompts()

                try:
                    new_image_path = do_ocr(
                        time_stamp, proba_id, proba_idc, tc_result_idc, tc_result_id, list_of_var_id, image, processed_image, processed_black,
                        ocr_prompts, image_path, self.excel_path, self.wb_target, model_cnn_status, model_dnn_status, model_ann_status,
                        model_resnet_status, model_rfg_status, cropped
                    )

                    # if checksheet_status:
                        # Move processed image to storage
                    if os.path.exists(new_image_path):
                        # self.save_status = True
                        self.move_processed_file(new_image_path)
                    # else:
                    #     self.save_status = False
                except Exception as e:
                    now = datetime.now()
                    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
                    traceback.print_exc()
                    logger.error(f"Error doing ocr for {image_path}: {str(e)}")
                    self.log_error_to_file(timestamp, e)
                    self.save_excel()
                    # self.save_status = True
                    filename = self.name_file
                    self.create_bat_file(filename)
            else:
                try:
                    # Allocate result based on checksheet ID
                    time_stamp = datetime.today().strftime("%H-%M-%S")
                    directory, filename = os.path.split(image_path)
                    new_filename = f"checksheet_re-scan_{time_stamp}.jpg"
                    new_path = os.path.join(directory, new_filename)
                    os.rename(image_path, new_path)
                    self.move_processed_file(new_path)
                except Exception as e:
                    now = datetime.now()
                    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
                    logger.error(f"Error processing file {image_path}: {str(e)}")
                    self.log_error_to_file(timestamp, e)
                    self.save_excel()
                    filename = self.name_file
                    # self.save_status = True
                    self.create_bat_file(filename)
                finally:
                    self.currently_processing = True
                    self.check_folder()
        except Exception as e:
            now = datetime.now()
            timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
            logger.error(f"Error processing file {image_path}: {str(e)}")
            self.log_error_to_file(timestamp, e)
            self.save_excel()
            filename = self.name_file
            self.create_bat_file(filename)
        finally:
            self.currently_processing = False
            self.check_folder_and_save()

    def delete_processed_file(self, image_path):
        """Delete processed file from scan folder."""
        try:
            if os.path.exists(image_path):
                os.remove(image_path)
                logger.info(f"Deleted processed file: {image_path}")
            else:
                logger.warning(f"Tried to delete file but it does not exist: {image_path}")

            # After deleting file, check scan folder
            self.check_folder_and_save()
        except Exception as e:
            logger.error(f"Error deleting file {image_path}: {str(e)}")

    def move_processed_file(self, image_path):
        """Move processed file to result folder and keep track of it."""
        try:
            filename = os.path.basename(image_path)
            destination = os.path.join(self.temp_folder, f"SCANNED_IMAGE/{filename}")
            
            # Ensure unique filename if file exists
            counter = 1
            while os.path.exists(destination):
                name, ext = os.path.splitext(filename)
                destination = os.path.join(self.temp_folder, f"SCANNED_IMAGE/{name}_{counter}{ext}")
                counter += 1
            
            shutil.move(image_path, destination)
            self.processed_files.append(destination)
            logger.info(f"Moved processed file to: {destination}")
            
            # After deleting file, check scan folder
            self.check_folder_and_save()
        except Exception as e:
            logger.error(f"Error moving file {image_path}: {str(e)}")

    def check_folder(self):
        """Check if scan folder is empty, only if paper has been scanned"""
        try:
            files_in_scan = []
            try:
                files_in_scan = [f for f in os.listdir(self.scan_folder) 
                               if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
                self.files_in_scan = files_in_scan
            except Exception as e:
                logger.error(f"Error reading scan folder: {str(e)}")
                return
            
            logger.debug(f"Files in scan folder: {files_in_scan}")
            
            # If there is no image and no process
            if not self.currently_processing and not files_in_scan:
                logger.info("No more files in scan folder.")
                
                # Empty processed file list
                self.processed_files = []
        except Exception as e:
            logger.error(f"Error in check_folder: {str(e)}")

    def check_folder_and_save(self):
        """Check if scan folder is empty and save excel if needed."""
        try:
            # Gunakan listdir dengan error handling
            files_in_scan = []
            try:
                files_in_scan = [f for f in os.listdir(self.scan_folder) 
                               if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
                self.files_in_scan = files_in_scan
            except Exception as e:
                logger.error(f"Error reading scan folder: {str(e)}")
                return
            
            logger.debug(f"Files in scan folder: {files_in_scan}")
            
            # Jika tidak ada file gambar dan tidak sedang memproses
            if not self.currently_processing and not files_in_scan:
                logger.info("No more files in scan folder. Saving excel...")
                self.save_excel()
                filename = self.name_file
                self.create_bat_file(filename)

                # Empty processed file list
                self.processed_files = []
                logger.info("Excel saved. Ready for new files.")
        except Exception as e:
            logger.error(f"Error in check_folder_and_save: {str(e)}")

    def save_excel(self):
        """Save the excel file with error handling."""
        try:
            # if self.save_status or not self.files_in_scan:
            print("Saving excel...")
            # Create a new workbook to save while keeping the original open
            temp_excel_path = self.excel_path + ".temp"
            print(temp_excel_path)
            self.wb_target.save(temp_excel_path)
            
            # Close the original workbook
            self.wb_target.close()

            # Replace the original with the new saved file
            shutil.move(temp_excel_path, self.excel_path)
            os.chmod(self.excel_path, 0o666)

            if os.path.exists(temp_excel_path):
                try:
                    os.remove(temp_excel_path)
                    logger.info(f"Deleted temporary Excel file: {temp_excel_path}")
                    print(f"Deleted temporary Excel file: {temp_excel_path}")
                except Exception as e:
                    logger.error(f"Error deleting temporary Excel file {temp_excel_path}: {e}")
                    print(f"Error deleting temporary Excel file {temp_excel_path}: {e}")

            directory, file_name = os.path.split(self.excel_path)
            self.rename = "done_" + file_name
            self.name_file, self.ext = os.path.splitext(self.rename)
            rename_path = os.path.join(RESULT_FOLDER, self.rename)
            print("self excel path", self.excel_path)
            print(rename_path)
            shutil.move(self.excel_path, rename_path)
            print(f"Excel renamed as {self.rename}")

            logger.info("Excel file saved  successfully.")
            print("Excel file saved successfully.")

            current_date = datetime.today().strftime("%Y-%m-%d_%H-%M-%S")
            excel_name = f"data_{current_date}"
            self.excel_path = os.path.join(self.excel_temp_folder, f"{excel_name}.xlsx")

            if not os.path.exists(self.excel_path):
                logger.info("Copying excel template...")
                shutil.copyfile(EXCEL_TEMPLATE_PATH, self.excel_path)

            # Reopen the workbook of the new path for further processing
            self.wb_target = openpyxl.load_workbook(self.excel_path)
            
            logger.info("Excel file copied and loaded successfully.")
            print("Excel file copied and loaded successfully.")

        except Exception as e:
            logger.error(f"Error saving excel file: {str(e)}")
            # Try to reopen the workbook if save failed
            try:
                self.wb_target = openpyxl.load_workbook(self.excel_path)
            except:
                logger.error("Failed to reload workbook. System may need restart.")

    def create_bat_file(self, excel_name):
        try:    
            filepath = os.path.join(RESULT_FOLDER, f"{excel_name}.bat")
            with open(filepath, "w") as f:
                f.write("@echo off\n")
                f.write("set SSH_USER=seizou-admin\n")
                f.write("set SSH_HOST=10.1.10.25\n")
                f.write("set SSH_PORT=22\n")
                f.write(f"set \"script_name={excel_name}\"\n") 

                f.write("set REMOTE_EXCEL_TO_JSON_PATH=/home/seizou-admin/PROJ/SEIZOUBU/OCR/HONSHA/SYSTEM/Program/excel_to_json.py\n")
                f.write("set REMOTE_MAPPING_SCRIPT_PATH=/home/seizou-admin/PROJ/SEIZOUBU/OCR/HONSHA/SYSTEM/Program/mapping_json_to_mariadb.py\n")

                f.write("set \"REMOTE_COMMANDS=python3 %REMOTE_EXCEL_TO_JSON_PATH% \"%script_name%\" && python3 %REMOTE_MAPPING_SCRIPT_PATH%\"\n")

                f.write("echo Attempting to connect to %SSH_USER%@%SSH_HOST% and run commands...\n")
                f.write("ssh -p %SSH_PORT% %SSH_USER%@%SSH_HOST% \"%REMOTE_COMMANDS%\"\n") 

                f.write("IF %ERRORLEVEL% NEQ 0 (\n")
                f.write("    echo.\n")
                f.write("    echo ERROR: SSH connection or remote command execution failed.\n")
                f.write("    echo Exit Code: %ERRORLEVEL%\n")
                f.write(") ELSE (\n")
                f.write("    echo.\n")
                f.write("    echo SSH connection and remote commands completed successfully.\n")
                f.write(")\n")
                f.write("pause")
            print(f"Successfully created batch file: {filepath}") 

        except Exception as e:
            logger.error(f"Error creating bat file: {e}") 

    def log_error_to_file(self, timestamp, exception):
        filename = f"error_{timestamp}.txt"
        path = os.path.join(RESULT_FOLDER, filename)
        try:
            # os.makedirs(RESULT_FOLDER, exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                f.write("❌ ERROR OCCURRED\n")
                f.write(f"Time: {timestamp}\n")
                f.write(f"Error: {exception}\n")
                f.write("Message: Please try scanning the checksheet again or contact admin at xxxx.\n")

            print(f"[LOG] File error saved in {path}")
        except Exception as log_e:
            print(f"[ERROR] Failed to save log error file: {log_e}")

    def wait_for_file(self, image_path, timeout=5):
        # waiting for file to be provided and copied
        start_time = time.time()
        while time.time() - start_time < timeout:
            if os.path.exists(image_path) and os.path.getsize(image_path) > 0:
                logger.info(f"File {image_path} is ready for processing.")
                return True
            time.sleep(0.5) # wait for 0.5 sec before rechecking
        logger.warning(f"File {image_path} not found or empty after {timeout} seconds.")
        return False

    def start_watchdog(self):
        """Start the watchdog observer."""
        self.executor = ThreadPoolExecutor(max_workers=1)
        event_handler = WatchdogHandler(self)
        self.observer = Observer()
        self.observer.schedule(event_handler, self.scan_folder, recursive=False)
        self.observer.start()
        logger.info("Watching for new files in scan folder...")
        print("Watching for new files in scan folder...")

    def stop_watchdog(self):
        """Stop the watchdog observer."""
        if self.observer:
            self.observer.stop()
            self.observer.join()
        if self.executor:
            self.executor.shutdown(wait=True)

class WatchdogHandler(FileSystemEventHandler):
    def __init__(self, file_processor):
        self.file_processor = file_processor

    def on_created(self, event):
        """Event ketika file baru dibuat di folder"""
        if event.is_directory:
            return  # Abaikan direktori, hanya file yang diproses

        if event.src_path.endswith(('.jpg', '.jpeg', '.png')):  # Memproses gambar
            logger.info(f"New file detected: {event.src_path}")
            image_path = os.path.normpath(event.src_path)
            print(image_path)

            if self.file_processor.wait_for_file(image_path):  # Pastikan file sudah siap
                self.file_processor.executor.submit(self.file_processor.process_new_file, image_path)
                    
if __name__ == "__main__":
    path_model_activate = PATH_MODEL_ACTIVATE
    with open(path_model_activate) as f:
        active_models = json.load(f)
    global model_rfg_status, model_ann_status, model_cnn_status, model_dnn_status, model_resnet_status
    model_rfg_status = active_models.get("model_rfg")
    model_ann_status = active_models.get("model_ann")
    model_dnn_status = active_models.get("model_dnn")
    model_cnn_status = active_models.get("model_cnn")
    model_resnet_status = active_models.get("model_resnet")
    
    processor = FileProcessor()
    processor.start_watchdog()
    processor.start_time = time.time
    temp_excel_path = processor.excel_path + ".temp"
    
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        processor.stop_watchdog()
        logger.info("Shutting down system...")
        print("Shutting down system...")
        
        # Final save before exiting
        filename = processor.name_file
        processor.create_bat_file(filename)
        print("Excel file saved and reloaded successfully.")
    finally:
        # Delete the temporary Excel file if it exists
        if os.path.exists(temp_excel_path):
            try:
                os.remove(temp_excel_path)
                logger.info(f"Deleted temporary Excel file: {temp_excel_path}")
                print(f"Deleted temporary Excel file: {temp_excel_path}")
            except Exception as e:
                logger.error(f"Error deleting temporary Excel file {temp_excel_path}: {e}")
                print(f"Error deleting temporary Excel file {temp_excel_path}: {e}")

        total_time = time.time() - processor.start_time()
        logger.info(f"\nProgram finished. Total execution time: {total_time:.2f} seconds.")
        print(f"\nProgram finished. Total execution time: {total_time:.2f} seconds.")